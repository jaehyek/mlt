# _*_ coding: utf-8 _*_

'''
Created on 2012. 7. 6.

@author: jaehyek
'''

import os , argparse
# from Crypto.Util.RFC1751 import wordlist




# variables.
diroutput = "."
outputtitle = "dwt_app_info"
outputexcel = outputtitle + ".xlsx"
outputcsv = outputtitle + ".csv"
outputexcelsheetname = outputtitle

boolFormatCSV = False


if __name__ == '__main__':
    pass

'''
usage : dwt_app_info.py  -o c:\temp_mlt 
'''

cmdlineopt = argparse.ArgumentParser(description='get the dwt_app_info  from MLT Server 10.185.135.57')
cmdlineopt.add_argument('-o', action="store", dest="diroutput", default='.',  help='output directory to put. default = . sample: c:\\temp_mlt' )
cmdlineopt.add_argument('-csv', action="store_true", dest="boolFormatCSV", default=False,  help='indicate the output format  default = False' )


cmdlineresults = cmdlineopt.parse_args()

diroutput = cmdlineresults.diroutput
boolFormatCSV = cmdlineresults.boolFormatCSV



if len(diroutput) == 0 :
    diroutput = "."
    

statement = "select * from dwt_app_info  "


    
diroutput = os.path.abspath(diroutput)
if not os.path.exists(diroutput):
    os.makedirs(diroutput)

''' goto the diroutput dir to work '''
os.chdir(diroutput)




import pymssql
conn = pymssql.connect(host='10.185.135.57', user='MLT_INF', password='inf_P@ssword', database='LGMLTPM', charset="UTF-8" , as_dict=True)
cur = conn.cursor()

cur.execute(statement)
results = cur.fetchall()
print "Total count = %s " % (len(results))

conn.close()

'''
keys = results[0].keys()
for key in keys:
    if type(key) == type("as"):
        print "'" + key + "', " ,

exit()
'''


''' field name '''

fieldnames = [
'app_name',
'category_cd',
'creation_id',
'preloaded_yn',
'pkg_name',
'etl_except_yn',
'last_update_id',
'local_app_name'
]



def saveExcelFormat (results ):
    print "Saving results in Excel format "

    ''' open the xls file '''
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.create_sheet(0) # insert at the end (default)
    ws.title = outputexcelsheetname
    rindex = 0 
    cindex = 0
     
    ''' first, insert field name to row0 '''
    for fieldname in fieldnames :
        ws.cell(row = rindex, column = cindex).value = fieldname
        cindex += 1 


    ''' print Symptom Info to excel file '''    
    rindex = 1 
    cindex = 0


    for record in results:
        for fieldname in fieldnames :
            ws.cell(row = rindex, column = cindex).value = record[fieldname]
            cindex += 1  
        rindex += 1
        cindex = 0 
        print  ".",

    print " "
    curdir = os.path.abspath(os.curdir)

    print "Excel Saving : " + curdir + "\\" +  outputexcel
    wb.save( outputexcel)

def saveCSVForamt  (results ):
    print "Saving results in CSV format "
    f = open ( outputcsv , "w" )
    ''' write the head line  '''
    str = ""
    for fieldname in fieldnames :
        str += fieldname + ", "
    str += "\n"
    f.write ( str )
    
    for record in  results : 
        print ".", 
        str = ""
        for fieldname in fieldnames :
            temp = record.get(fieldname, "")
            if temp == None :
                temp = ""
            str += temp.encode("utf-8") + ", " 
        str += "\n"
        f.write ( str )
    f.close()
    print " "

if boolFormatCSV :
    saveCSVForamt  (results )
else:
    saveExcelFormat (results )
    
print "Processing End ....."
exit()





