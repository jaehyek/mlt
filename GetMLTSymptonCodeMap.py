# _*_ coding: utf-8 _*_

'''
Created on 2012. 7. 6.

@author: jaehyek
'''

import os , argparse
# from Crypto.Util.RFC1751 import wordlist




# variables.
diroutput = "."
outputtitle = "GetMLTSymptonCodeMap"
outputexcel = outputtitle + ".xlsx"
outputexcelsheetname = outputtitle

if __name__ == '__main__':
    pass

'''
usage : GetMLTSymptonCodeMap.py  -o c:\temp_mlt 
'''

cmdlineopt = argparse.ArgumentParser(description='get the GetMLTSymptonCodeMap  from MLT Server 10.185.135.57')
cmdlineopt.add_argument('-o', action="store", dest="diroutput", default='.',  help='output directory to put. default = . sample: c:\\temp_mlt' )


cmdlineresults = cmdlineopt.parse_args()

diroutput = cmdlineresults.diroutput



if len(diroutput) == 0 :
    diroutput = "."
    

statement = "select * from dwt_symptom_m"


    
diroutput = os.path.abspath(diroutput)
if not os.path.exists(diroutput):
    os.makedirs(diroutput)

''' goto the diroutput dir to work '''
os.chdir(diroutput)




import pymssql
conn = pymssql.connect(host='10.185.135.57', user='MLT_INF', password='inf_P@ssword', database='LGMLTPM', charset="UTF-8" , as_dict=True)
cur = conn.cursor()

cur.execute(statement)
results = cur.fetchall()
print "Total count = %s " % (len(results))

conn.close()

'''
keys = results[0].keys()
for key in keys:
    if type(key) == type("as"):
        print "'" + key + "', " ,

exit()
'''

''' open the xls file '''
from openpyxl import Workbook

''' field name '''
# fieldnames = [ 'symptom_cd',  'symptom_description',  'section_cd','section_description',  'cause_cd',  'cause_description',  'mapping_description', 'creation_dt',  'mapping_defect_id',   'creation_id',    'source_system',  'last_update_dt',  'last_update_id' ]
fieldnames = [ 'symptom_key_code','source_system','creation_dt','section_cd','symptom_desc','corporation_code','mapping_defect_id','symptom_cd','section_desc' ]





wb = Workbook()
ws = wb.create_sheet(0) # insert at the end (default)
ws.title = outputexcelsheetname
rindex = 0 
cindex = 0
 
''' first, insert field name to row0 '''
for fieldname in fieldnames :
    ws.cell(row = rindex, column = cindex).value = fieldname
    cindex += 1 


''' print Symptom Info to excel file '''    
rindex = 1 
cindex = 0


for record in results:
    for fieldname in fieldnames :
        ws.cell(row = rindex, column = cindex).value = record[fieldname]
        cindex += 1  
    rindex += 1
    cindex = 0 
    print  ".",

print " "
curdir = os.path.abspath(os.curdir)

print "Excel Saving : " + curdir + "\\" +  outputexcel
wb.save( outputexcel)



print "Processing End ....."
exit()





